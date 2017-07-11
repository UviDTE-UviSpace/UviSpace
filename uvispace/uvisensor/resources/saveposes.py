#!/usr/bin/env python
"""Auxiliary program for manipulate data of poses and time.

This module allows:
-Read data of poses with their respective time, from a spreadsheet.
-Analyze data by calculating differential values of position, time and
 lenght displaced and average speed of UGV
-Save data in spreadsheet and textfile.
-Save final data in master sheet.
"""
# Standard libraries
import glob
import numpy as np
import os
import re
from scipy import stats
import sys
import time
# Excel read/write library
import openpyxl

class DataAnalyzer(object):
    """Receives time and poses of matrix to filter and analyze.

    Different time and position values are calculated between two data
    points. From these, the linear average speed and the angular average
    speed are calculated.

    :param data: Matrix of floats64 with M data.
    :type data: numpy.array float64 (shape=Mx4).
    :param analyzed_data: Matrix with M analyzed data after calculating.
     differential values and relative linear and angular speeds.
    :type analyzed_data: numpy.array float64 (shape=Mx11).
    :param float64 avg_speed: average linear speed of UGV.
    :param float64 avg_ang_speed: average angular speed of UGV.
    """
    def __init__(self, data=np.zeros((1,4)), analyzed_data=np.zeros((1,11)),
                 avg_speed=0, avg_ang_speed=0):
        self._raw_data = data
        self._analyzed_data = analyzed_data
        self._avg_speed = avg_speed
        self._avg_ang_speed = avg_ang_speed

    def upload_data(self, data):
        """Update data matrix with time and poses.

        :param data: Matrix of floats64 with M data.
        :type data: numpy.array float64 (shape=Mx4).
        :return: updated data.
        :rtype: numpy.array float64 (shape=Mx4).
        """
        self._raw_data = data
        return self._raw_data

    def remove_repeated_poses(self):
        """Remove repeated pose values.

        When the camera does not detect a triangle, the pose value is
        not updated, remaining the current pose as the previous pose.
        With this method, these values are eliminated.

        :return: data without repeated values.
        :rtype: numpy.array float64 (shape=Mx4).
        """
        # New data is correct when at least one value of pose are different.
        rows = self._raw_data.shape[0]
        # Initialization data arrays.
        last_data = np.array([0, 0, 0, 0]).astype(np.float64)
        filtered_data = np.array([0, 0, 0, 0]).astype(np.float64)
        for row in range(rows):
            # New dara read array.
            new_data = np.copy(self._raw_data[row, :])
            different_values = np.any(new_data[1:4] != last_data[1:4])
            if different_values:
                last_data = np.copy(new_data)
                if row == 0:
                    # Initialization filtered_data array.
                    filtered_data = np.copy(new_data)
                else:
                    filtered_data = np.vstack([filtered_data, new_data])
        self._raw_data = np.copy(filtered_data)
        return filtered_data

    def remove_stop_poses(self):
        """Remove stop pose values.

        This method eliminates the initial and final values in which the
        UGV is stopped.

        :return: data without stop values.
        :rtype: numpy.array float64 (shape=Mx4).
        """
        rows = self._raw_data.shape[0]
        # Initialization first and last rows witout poses with stopped UGV.
        row_upper = 0
        row_lower = rows
        # Initial repeated data calculation.
        mode_upper = stats.mode(self._raw_data[0:20, 1:3])[0]
        mode_rows = np.all(self._raw_data[:, 1:3]==mode_upper, axis=1)
        indexes = np.where(mode_rows)[0]
        # If there are not initial values of stopped UGV is not updated.
        if indexes.shape[0] > 0:
            row_upper = indexes.max()
        # Final repeated data calculation.
        mode_lower = stats.mode(self._raw_data[(rows-20):rows, 1:3])[0]
        mode_rows = np.all(self._raw_data[:, 1:3]==mode_lower, axis=1)
        indexes = np.where(mode_rows)[0]
        # If there are not final values of stopped UGV is not updated.
        if indexes.shape[0] > 0:
            row_lower = indexes.min() + 1
        # UGV data in motion.
        clipped_data = self._raw_data[row_upper:row_lower, :]
        self._raw_data = np.copy(clipped_data)
        return clipped_data

    def get_diff_data(self):
        """Get differential data and relative linear and angular speed.

        Difference data are obtained between two consecutive samples,
        and linear and angular speeds are calculated from them.

        It also obtain the absolute linear and angular velocity of the
        experiment.

        :returns: [formatted_data, avg_speed, avg_ang_speed]
          * *formatted_data* data with differential values and angular
            and linear speeds.
          * *avg_speed* average linear speed of UGV.
          * *avg_ang_speed* average angular speed of UGV.
        :rtype: [numpy.array float64 (shape=Mx11), float64, float64]
        """
        # Calculate differential values.
        work_data = np.copy(self._raw_data)
        rows = self._raw_data.shape[0]
        # First sample, time zero.
        work_data[:, 0] -= work_data[0, 0]
        # Differential data matrix: current data minus previous data.
        diff_data = np.zeros_like(work_data)
        diff_data[1:rows, :] = work_data[1:rows, :] - work_data[0:(rows-1), :]
        # first_row = np.zeros((1,4))
        # diff_data = np.vstack([first_row, diff_data])
        # Vector differential length displaced.
        diff_length = np.sqrt(diff_data[:, 1]**2 + diff_data[:, 2]**2)
        # The direction is negative when the angular speed and vehicle angle
        # difference is bigger than pi/2 (90 degrees).
        speed_angles = np.arctan2(diff_data[:, 2], diff_data[:, 1])
        #import pdb; pdb.set_trace()
        sign_spd = np.ones([(rows), 1])
        sign_spd[np.abs(work_data[:, 3] - speed_angles) > (np.pi/2)] *= -1
        diff_speed = sign_spd[1:, 0] * 1000 * diff_length[1:]/diff_data[1:, 0]
        diff_speed = np.hstack([0, diff_speed])
        # Vector differential angular speed.
        diff_angl_speed = np.zeros(rows)
        diff_angl_speed[1:rows] = 1000 * diff_data[1:rows, 3] / diff_data[1:rows, 0]
        # Complete differential data matrix with new data.
        # diff_data = np.hstack([np.zeros(rows, 1), diff_data])
        diff_data = np.insert(diff_data, 4, diff_length, axis=1)
        diff_data = np.insert(diff_data, 5, diff_speed, axis=1)
        diff_data = np.insert(diff_data, 6, diff_angl_speed, axis=1)
        # Complete data matrix with differential_data.
        work_data = np.hstack([work_data, diff_data])
        # Average speed and average angular speed.
        sum_data = diff_data.sum(axis=0)
        length = np.sqrt(((work_data[rows-1,1] - work_data [0, 1])**2)
                           + ((work_data[rows-1,2] - work_data [0, 2])**2))
        sum_data = diff_data.sum(axis=0)
        avg_speed = np.round((1000*length/sum_data[0]), 2)
        avg_ang_speed = np.round(((1000*(work_data[rows-1, 3]
                                   - work_data [0, 3]))/sum_data[0]), 2)
        formatted_data = np.round(work_data, 2)
        self._analyzed_data = np.copy(formatted_data)
        self._avg_speed = avg_speed
        self._avg_ang_speed = avg_ang_speed
        return (formatted_data, avg_speed, avg_ang_speed)

def read_data(name_spreadsheet="31_05_2017_201-L255-R255b.xlsx"):
    """It allows to read poses and time of spreadsheet.

    These data are stored in the "data" matrix.

    Each group of data is composed to 4 elements about the location of
    the UGV at a given time: time, position "x", position "y" and angle
    "theta".

    :param str name_spreadsheet: name of spreadsheet that contain the
    data to be read.
    :return: matrix dimentions Mx4 with read data. M is the number of
    rows corresponding to the number of data read.
    :rtype: numpy.array(shape=Mx4)
    """
    # Open spreadsheet
    try:
        wb = openpyxl.load_workbook(name_spreadsheet)
    except IOError:
        wb = openpyxl.Workbook()
    ws = wb.active
    # Initialization of matrixes for data.
    data = np.array([0, 0, 0, 0]).astype(np.float64)
    new_data = np.array([0, 0, 0, 0]).astype(np.float64)
    # Rows 1 and 2 are for the file name, the next three rows describe the
    # experiment parameters, and the 6 is the header in any type file.
    # Start reading in row 7.
    row = 7
    # Number of columns in the matrix.
    cols = data.shape[0]
    # Loop for reading data.
    data_in_this_row = True
    while data_in_this_row:
        element = ws.cell(column=1, row=row).value
        # "Sum differerential data:" is the value of the next row to the last
        # data.
        if element == 'Sum differential data:':
            data_in_this_row = False
        else:
            # Reading data.
            for col in range (0, cols):
                element = ws.cell(column=col+1, row=row).value
                new_data[col] = element
            if row == 7:
                # Substitution of row of zeros by first row of data read.
                data = np.copy(new_data)
            else:
                data = np.vstack([data, new_data])
            row +=1
    rounded_data = np.round(data, 2)
    return rounded_data

"""Dictionary with different formats to apply to the cells when writes
the data in a spreadsheet.
"""
cell_formats = {
    # Types of text alignment.
    'center_al': openpyxl.styles.Alignment(horizontal='center',
                                           vertical='center'),
    'right_al': openpyxl.styles.Alignment(horizontal='right',
                                          vertical='center'),
    # Types of font.
    'title_ft': openpyxl.styles.Font(color='FFFFFFFF', size=20, bold=True),
    'white_ft': openpyxl.styles.Font(color='FFFFFFFF', bold=True),
    # Types of border.
    'thick_bd': openpyxl.styles.Border(
                                top=openpyxl.styles.Side(border_style='thick',
                                                         color='FF4143CA'),
                                bottom=openpyxl.styles.Side(
                                       border_style='thick', color='FF4143CA')),
    'thin_bd': openpyxl.styles.Border(
                                   top=openpyxl.styles.Side(border_style='thin',
                                                            color='FF83C6D6'),
                                bottom=openpyxl.styles.Side(border_style='thin',
                                                            color='FF83C6D6')),
    # Types of fill.
    'blue_fill': openpyxl.styles.PatternFill(fill_type='solid',
                                             start_color='FF2F79E6',
                                             end_color='FF2F79E6'),
    'skyblue_fill': openpyxl.styles.PatternFill(fill_type='solid',
                                                start_color='FFDBF4FA',
                                                end_color='FFDBF4FA'),
    'white_fill': openpyxl.styles.PatternFill(fill_type='solid',
                                              start_color='FFFFFFFF',
                                              end_color='FFFFFFFF')
}
class DataSaver(DataAnalyzer):

    def __init__(self, data=np.zeros((1,4)), analyzed_data=np.zeros((1,11)),
                 avg_speed=0, avg_ang_speed=0):
        self._filename = None
        self._sp_left = 0
        self._sp_right = 0
        self._exp_conditions = ""
        self._data_to_save = 0
        self._header = ""
        self._save_analyzed = False

    def upload_setpoints_save(self, sp_left, sp_right):

        self._sp_left = sp_left
        self._sp_right = sp_right

        return sp_left, sp_right

    def upload_filename_save(self, filename):

        self._filename = filename

        return filename

    def save2data(self, save_analyzed=False):
        self._save_analyzed = save_analyzed
        if self._save_analyzed:
            work_data = np.copy(self._analyzed_data)
            # Header construction.
            self._header = np.array(['Time', 'Pos x', 'Pos y', 'Angle',
                                    'Diff Time', 'Diff Posx', 'Diff Posy',
                                    'Diff Angl', 'Diff Leng', 'Rel Speed',
                                    'Rel AnSpd'])
        else:
            work_data = np.copy(self._raw_data)
            # Header construction.
            self._header = np.array(['Time', 'Pos x', 'Pos y', 'Angle'])
        #First sample, time zero.
        rows, cols = work_data.shape
        work_data[:, 0] = work_data[:, 0] - work_data[0, 0]
        self._data_to_save = np.copy(work_data)
        # Name of the output file for the poses historic values.
        if not self._filename:
            exist_file = glob.glob("datatemp/*.xlsx")
            exist_file.sort()
            index = len (exist_file)
            datestamp_file = "{}".format(time.strftime("%d_%m_%Y"))
            self._filename = "{}_{}-L{}-R{}".format(datestamp_file, (index+1),
                            self._sp_left, self._sp_right)
            datestamp_text = '{}_{}'.format(datestamp_file, (index+1))
        name_txt = "datatemp/{}.txt".format(self._filename)
        #Header for numpy savetxt.
        header_numpy = ''
        cols = self._header.shape[0]
        for col in range (0, cols):
            element = self._header[col]
            element = '%9s' % (element)
            header_numpy = '{}{}\t'.format(header_numpy, element)
        #Call to save data in textfile.
        np.savetxt(name_txt, work_data, delimiter='\t', fmt='%9.2f',
                   header=header_numpy, comments='')
        #Experiment conditions.
        self._exp_conditions = (" -Use camera 3\n -Position initial experiment forward: "
                          "right rear wheel profile (-1400, -600), rear axis UGV in "
                          "axis y, in -1800 x\n -Time: 3 seconds")
        self.write_spreadsheet()
        return work_data

    def write_spreadsheet(self):
        """
        Receives poses and time, and saves them in a spreadsheet.

        :param header: contains header to save in spreadsheet.
        :param data: contains data to save in spreadsheet.
        :param filename: name of spreadsheet where the data will be saved.
        :param exp_conditions: contains string with experiment description.
        :param save_master: boolean with True for save data in masterfile.
        """
        name_spreadsheet = "datatemp/{}.xlsx".format(self._filename)
        try:
            wb = openpyxl.load_workbook(name_spreadsheet)
        except:
            wb = openpyxl.Workbook()
        ws = wb.active
        #Spreadsheet title.
        ws.merge_cells('A1:K2')
        ws.cell('A1').value = self._filename
        ws.cell('A1').alignment = cell_formats['center_al']
        ws.cell('A1').font = cell_formats['title_ft']
        ws.cell('A1').fill = cell_formats['blue_fill']
        #Experiment conditions.
        ws.merge_cells('A3:K5')
        ws.cell('A3').value = self._exp_conditions
        #Freeze header
        my_cell = ws['B7']
        ws.freeze_panes = my_cell
        #Write in spreadsheet the headboard.
        rows, cols = self._data_to_save.shape
        for col in range (0, cols):
            ws.cell(column=col+1, row=6, value=self._header[col])
            ws.cell(column=col+1, row=6).alignment = cell_formats['right_al']
            ws.cell(column=col+1, row=6).font = cell_formats['white_ft']
            ws.cell(column=col+1, row=6).fill = cell_formats['blue_fill']
            ws.cell(column=col+1, row=6).border = cell_formats['thick_bd']
        #Write in spreadsheet the data.
        for row in range(0, rows):
            for col in range(0, cols):
                element = float(self._data_to_save[row,col])
                ws.cell(column=col+1, row=row+7, value=element).number_format = '0.00'
                if row % 2 != 0:
                    ws.cell(column=col+1, row=row+7).fill = cell_formats[
                                                                 'skyblue_fill']
                else:
                    ws.cell(column=col+1, row=row+7).fill = cell_formats[
                                                                   'white_fill']
                my_cell = ws.cell(column=col+1, row=row+7)
                ws.column_dimensions[my_cell.column].width = 10
        #Write in spreadsheet the name of statistics.
        if self._save_analyzed:
            for row in range(rows+7, rows+11):
                ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
                ws.cell(column=1, row=row).alignment = cell_formats['right_al']
                ws.cell(column=1, row=row).fill = cell_formats['blue_fill']
                ws.cell(column=1, row=row).font = cell_formats['white_ft']
            ws.cell(column=1, row=rows+7, value='Sum differential data:')
            ws.cell(column=1, row=rows+8, value='Mean of differential data:')
            ws.cell(column=1, row=rows+9, value='Variance differential data:')
            ws.cell(column=1, row=rows+10, value='Std deviation differential data:')
            ws.cell(column=8, row=rows+11, value='Linear Relative Speed:')
            ws.cell(column=8, row=rows+12, value='Angular Relative Speed:')
            ws.merge_cells(start_row=rows+11,start_column=8,end_row=rows+11,end_column=10)
            ws.merge_cells(start_row=rows+12,start_column=8,end_row=rows+12,end_column=10)
            ##Write and calculate in spreadsheet the statistics.
            for col in range(5, cols+1):
                letter_range = openpyxl.utils.get_column_letter(col)
                start_range = '{}{}'.format(letter_range, 7)
                end_range = '{}{}'.format(letter_range, rows+5)
                interval = '{}:{}'.format(start_range,end_range)
                ws.cell(column=col, row=rows+7, value= '=SUM({})\n'.format(interval))
                ws.cell(column=col, row=rows+8, value= '=AVERAGE({})\n'.format(interval))
                ws.cell(column=col, row=rows+9, value= '=VAR({})\n'.format(interval))
                ws.cell(column=col, row=rows+10, value= '=STDEV({})\n'.format(interval))
                for row in range(rows+7, rows+13):
                    ws.cell(column=col, row=row).number_format = '0.00'
                    ws.cell(column=col, row=row).font = cell_formats['white_ft']
                    ws.cell(column=col, row=row).fill = cell_formats['blue_fill']
                    ws.cell(column=col, row=row).alignment = cell_formats['right_al']
            ws.cell(column=11, row=rows+11, value= '=1000*SQRT(((B{rows}-B7)^2)+'
                    '(((C{rows}-C7)^2)))/E{rows2}\n'.format(rows=rows+5, rows2=rows+7))
            ws.cell(column=11, row=rows+12, value= '=1000*(D{rows}-D7)/'
                                        'E{rows2}\n'.format(rows=rows+5, rows2=rows+7))
        wb.save(name_spreadsheet)

        return name_spreadsheet

    def save2master_xlsx(data_master):
        """
        Average speed, setpoint left and right wheels are saved in the same spreadsheet.

        :param data_master: array that contains the average lineal speed, the
        average angular speed, the set point left, the set point right, and the
        name of datafile.
        """
        #Data search to save in masterspreadsheet.
        folder = '\'file:///home/jorge/UviSpace/uvispace/uvisensor/'
        name_sheet = '\'#$Sheet.'
        try:
            wb = openpyxl.load_workbook(data_master[3])
        except:
            wb = openpyxl.Workbook()
        ws = wb.active
        #Next empty row search.
        row = 7
        written_row = True
        while written_row:
            element = ws.cell(column=1, row=row).value
            if element == None:
                written_row = False
            else:
                row +=1
        avg_speed = folder + data_master[3] + name_sheet + 'K' + '{}'.format(row)
        row = row + 1
        avg_ang_speed = folder + data_master[3] + name_sheet + 'K' + '{}'.format(row)
        #Save data in masterspreadsheet.
        try:
            wb = openpyxl.load_workbook("datatemp/masterfile4.xlsx")
        except:
            wb = openpyxl.Workbook()
        ws = wb.active
        #Freeze header
        my_cell = ws['A2']
        ws.freeze_panes = my_cell
        #Next empty row search.
        row = 1
        written_row = True
        while written_row:
            element = ws.cell(column=1, row=row).value
            if element == None:
                written_row = False
            else:
                row +=1
        #Save data in empty row.
        ws.cell(column=1, row=row, value=data_master[0])
        ws.cell(column=2, row=row, value=data_master[1])
        ws.cell(column=3, row=row, value=data_master[2])
        ws.cell(column=4, row=row, value= '=INDIRECT(F{})\n'.format(row)).number_format = '0.00'
        ws.cell(column=5, row=row, value= '=INDIRECT(G{})\n'.format(row)).number_format = '0.00'
        ws.cell(column=6, row=row, value=avg_speed)
        ws.cell(column=7, row=row, value=avg_ang_speed)
        ws.cell(column=8, row=row, value="_")
        #Format data.
        for y in range (1, 8):
            my_cell = ws.cell(column=y, row=row)
            if y == 1:
                ws.column_dimensions[my_cell.column].width = 18
            else:
                ws.column_dimensions[my_cell.column].width = 10
            if row % 2 != 0:
                ws.cell(column=y, row=row).fill = cell_formats['skyblue_fill']
            else:
                ws.cell(column=y, row=row).fill = cell_formats['white_fill']
            if y < 6:
                ws.cell(column=y, row=row).alignment = cell_formats['right_al']
        wb.save("datatemp/masterfile4.xlsx")

    def save2master_txt(data_master):
        """
        Average speed, setpoint left and right wheels are saved in the same textfile.

        :param data_master: array that contains the average lineal speed, the
        average angular speed, the set point left, the set point right, and the
        name of datafile.
        """
        #TODO improve format
        cols = data_master.shape[0]
        for y in range (0, cols):
            value = data_master[y]
            if y == 0:
                #TODO use '.format' string construction style
                text = '%9s' % (value)
            else:
                value = float(value)
                value = '%9.2f' % (value)
                text = '{}\t\t{}'.format(text, value)
        text = '{}\n'.format(text)
        with open("datatemp/masterfile4.txt", 'a') as outfile:
            outfile.write(text)


def main():
    # o_exist_file = glob.glob("./datatemp/*.xlsx")
    # o_exist_file.sort()
    # o_index = len (o_exist_file)
    # names = [os.path.basename(x) for x in o_exist_file]
    # import pdb; pdb.set_trace()
    # for y in range (0, o_index-1):
    #     #import pdb; pdb.set_trace()
    #     print names[y]
    #     data_matrix = read_data(name_spreadsheet='datatemp/' + names[y])
    #
    # #numbers_filename = re.findall(r'\d+', name_spreadsheet)
    # #sp_left = int(numbers_filename[4])
    # #sp_right = int(numbers_filename[5])
    #
    #
    # #TODO Try, except correct value.
    # sp_left = input("Introduce value of sp_left between 0 and 255\n")
    # sp_right = input("Introduce value of sp_right between 0 and 255\n")
    data1 = read_data()
    treatment = DataSaver()
    treatment.upload_data(data1)
    treatment.get_diff_data()
    treatment.save2data(True)

if __name__ == '__main__':
    main()